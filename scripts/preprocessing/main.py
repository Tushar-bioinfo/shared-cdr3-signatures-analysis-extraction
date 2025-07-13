import pandas as pd
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# ---- CONFIGURATION ----
cancer_inputs = {
    "Kidney": "/Users/tusharsingh/Desktop/work/raw_data/IGH/run/kidney_tumor.csv",
    "Uterus": "/Users/tusharsingh/Desktop/work/raw_data/IGH/run/uterus_tumor.csv",
    "Lung_adenocarcinoma": "/Users/tusharsingh/Desktop/raw_data/TRB/Lung_AC_TRB.csv",
    "Lung_squamous": "/Users/tusharsingh/Desktop/raw_data/TRB/Lung_SCC_TRB.csv",
}
cdr3_column = "CDR3"
base_output_root = "/Users/tusharsingh/Desktop/motifs_may/Output/TRB/"
kmer_range = range(4,12)  # k = 4 to 11

# ---- FUNCTION TO EXTRACT KMERS ----
def extract_kmers(sequences, k):
    counter = Counter()
    for seq in sequences:
        if isinstance(seq, str) and len(seq) >= k:
            for i in range(len(seq) - k + 1):
                counter[seq[i:i + k]] += 1
    return counter

# ---- MAIN LOOP OVER K-MER LENGTHS ----
for kmer_size in kmer_range:
    print(f"\n🎯 Processing k-mer length: {kmer_size}")
    output_root = f"{base_output_root}_k{kmer_size}"
    os.makedirs(output_root, exist_ok=True)

    top_motifs = {}

    for label, csv_path in cancer_inputs.items():
        print(f"\n🔍 Processing {label} from {csv_path}")
        df = pd.read_csv(csv_path)
        cdr3_all = df[cdr3_column].dropna().tolist()
        cdr3_unique = list(set(cdr3_all))

        count_all = extract_kmers(cdr3_all, kmer_size)
        count_unique = extract_kmers(cdr3_unique, kmer_size)
        all_kmers = sorted(set(count_all) | set(count_unique))

        df_merged = pd.DataFrame([
            (kmer, count_all.get(kmer, 0), count_unique.get(kmer, 0))
            for kmer in all_kmers
        ], columns=[f"{kmer_size}-mer", "Count_All", "Count_Unique"])
        df_merged.sort_values("Count_All", ascending=False, inplace=True)

        # Store top 50 motif names only
        top_motifs[label] = df_merged.head(50)[f"{kmer_size}-mer"].tolist()

        # Save files to labeled output folder
        out_dir = Path(output_root) / f"{label}_output"
        out_dir.mkdir(exist_ok=True, parents=True)
        df_merged.to_csv(out_dir / f"{label}_kmer_summary.csv", index=False)
        print(f"✅ Saved: {out_dir / f'{label}_kmer_summary.csv'}")

    # ---- COMBINE TOP 50 MOTIF NAMES FOR OBSERVATION ----
    final_rows = []
    max_len = 50
    labels = list(top_motifs.keys())

    for i in range(max_len):
        row = []
        for label in labels:
            top_list = top_motifs.get(label, [])
            row.append(top_list[i] if i < len(top_list) else "")
            row.append("")  # spacer column
        final_rows.append(row)

    cols = []
    for label in labels:
        cols.append(label)
        cols.append(f"Spacer_{label}")

    obs_df = pd.DataFrame(final_rows, columns=cols)
    obs_csv_path = Path(output_root) / f"Top50_Motifs_Observation_Summary_k{kmer_size}.csv"
    obs_df.to_csv(obs_csv_path, index=False)
    print(f"\n📊 Summary CSV saved to: {obs_csv_path}")

    # ---- COLOR-CODED EXCEL EXPORT ----
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Top50_k{kmer_size}"

        ws.append(cols)

        color_palette = ["FFFF00", "FFCC99", "CCFFCC", "99CCFF", "FF9999", "CCCCFF", "FFB6C1", "D3FFCE"]
        color_map = defaultdict(lambda: "FFFFFF")
        color_idx = 0

        for row in final_rows:
            excel_row = []
            for i, motif in enumerate(row):
                if i % 2 == 0:
                    if motif and motif not in color_map:
                        color_map[motif] = color_palette[color_idx % len(color_palette)]
                        color_idx += 1
                    excel_row.append(motif)
                else:
                    excel_row.append("")
            ws.append(excel_row)

        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                if cell.value in color_map:
                    fill = PatternFill(start_color=color_map[cell.value], end_color=color_map[cell.value], fill_type="solid")
                    cell.fill = fill

        xlsx_path = Path(output_root) / f"Top50_Motifs_Color_k{kmer_size}.xlsx"
        wb.save(xlsx_path)
        print(f"🎨 Excel file saved: {xlsx_path}")
    except Exception as e:
        print(f"⚠️ Failed to write Excel file for k={kmer_size}: {e}")
